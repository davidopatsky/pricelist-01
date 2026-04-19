#!/usr/bin/env python3
"""Export pricelist.xlsx → pricelist.json (one-way, manual).

Schema-agnostic: reads ALL columns from the xlsx header row and passes
them through to pricelist.json. Adding, removing, or reordering columns
in the xlsx does NOT require changes to this script.

Only 3 things are validated:
  1. `sku` column must exist (required, unique per row)
  2. `category` column must exist (required, must be one of ALLOWED_CATEGORIES)
  3. `currency` column must exist (required)

Everything else is passed through as-is. Column order in the JSON output
matches the xlsx header order.

PRICING CONVENTION (see BLUEPRINT.md):
  standard_price=1 is a magic placeholder, not 1 CZK.
  Consumers must read standard_price + price_formula together.
  This script preserves both verbatim — no normalization.

Run:
    python3 scripts/export_json.py
    python3 scripts/export_json.py --dry-run
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'pricelist.xlsx'
OUT = ROOT / 'pricelist.json'

# Required columns — script will crash with a clear error if these are missing.
REQUIRED = {'sku', 'category', 'currency'}

ALLOWED_CATEGORIES = {'pergola', 'vypln', 'prislusenstvi', 'sluzba', 'elektro', 'priplatky'}

# Header aliases: common typos / legacy names → canonical key.
# Canonical keys are whatever pricelist.json currently uses.
HEADER_ALIASES = {
    'name-en':     'name_en',
    'name_cs':     'name',
    'nazev':       'name',
    'nazev-cz':    'name',
}

# Columns whose values should be treated as numbers when possible.
# Everything else is kept as text (string | null).
NUMERIC_COLUMNS = {'standard_price', 'cost', 'cost_percent', 'raynet_cost'}

def parse_args():
    return '--dry-run' in sys.argv[1:]

def normalize_value(raw, is_numeric):
    """Convert a raw cell value to a clean Python value."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if is_numeric:
            return int(raw) if float(raw).is_integer() else raw
        return raw  # preserve numbers even in text columns
    s = str(raw).strip()
    if s == '':
        return None
    if is_numeric:
        try:
            n = float(s)
            return int(n) if n.is_integer() else n
        except ValueError:
            return s  # keep as string (e.g. "matrix" in standard_price)
    return s

def read_xlsx():
    if not XLSX.exists():
        raise SystemExit(f'{XLSX} not found. Run scripts/init_xlsx.py first.')
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    # Read header row, apply aliases
    raw_headers = [str(c.value).strip() if c.value else None for c in ws[1]]
    columns = []  # list of (canonical_name, col_index)
    for i, h in enumerate(raw_headers):
        if not h:
            continue
        canonical = HEADER_ALIASES.get(h, h)
        columns.append((canonical, i))

    found_keys = {name for name, _ in columns}
    missing = REQUIRED - found_keys
    if missing:
        raise SystemExit(f'Missing required columns in xlsx header: {sorted(missing)}')

    items = []
    errors = []
    seen_skus = set()

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        entry = {}
        for name, idx in columns:
            raw = row[idx] if idx < len(row) else None
            entry[name] = normalize_value(raw, name in NUMERIC_COLUMNS)

        sku = entry.get('sku')
        if not sku:
            errors.append(f'row {r_idx}: sku is empty')
            continue
        if sku in seen_skus:
            errors.append(f'row {r_idx}: duplicate sku "{sku}"')
        seen_skus.add(sku)

        cat = entry.get('category')
        if not cat:
            errors.append(f'row {r_idx}: category is empty')
        elif cat not in ALLOWED_CATEGORIES:
            errors.append(f'row {r_idx}: category "{cat}" not in {sorted(ALLOWED_CATEGORIES)}')

        if not entry.get('currency'):
            errors.append(f'row {r_idx}: currency is empty')

        items.append(entry)

    return items, errors

def main():
    dry = parse_args()
    items, errors = read_xlsx()

    if errors:
        print('ERRORS:')
        for e in errors:
            print(f'  {e}')
        sys.exit(1)

    old = []
    if OUT.exists():
        old = json.loads(OUT.read_text())

    new_json = json.dumps(items, indent=2, ensure_ascii=False) + '\n'
    old_json = json.dumps(old, indent=2, ensure_ascii=False) + '\n'

    if new_json == old_json:
        print(f'No changes. {len(items)} items.')
        return

    added = [i['sku'] for i in items if i['sku'] not in {o['sku'] for o in old}]
    removed = [o['sku'] for o in old if o['sku'] not in {i['sku'] for i in items}]
    changed = 0
    old_by = {o['sku']: o for o in old}
    for i in items:
        if i['sku'] in old_by and i != old_by[i['sku']]:
            changed += 1

    # Detect new/removed columns
    old_keys = set()
    for o in old:
        old_keys.update(o.keys())
    new_keys = set()
    for i in items:
        new_keys.update(i.keys())
    added_cols = new_keys - old_keys
    removed_cols = old_keys - new_keys

    print('Changes:')
    print(f'  {len(added)} added: {added if added else "—"}')
    print(f'  {len(removed)} removed: {removed if removed else "—"}')
    print(f'  {changed} modified')
    if added_cols:
        print(f'  NEW COLUMNS: {sorted(added_cols)}')
    if removed_cols:
        print(f'  REMOVED COLUMNS: {sorted(removed_cols)}')

    if dry:
        print('(dry run — pricelist.json NOT written)')
        return

    OUT.write_text(new_json)
    print(f'Wrote {OUT} ({len(items)} items, {len(new_keys)} fields)')

if __name__ == '__main__':
    main()
