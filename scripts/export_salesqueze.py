#!/usr/bin/env python3
"""Export price matrices to individual xlsx files for SalesQueze.

Reads products/<category>/<sku>/prices.json for each SKU in SKUS and
writes export/salesqueze/<sku>.xlsx containing just two dimensions
(width × depth) and prices. Single-matrix products get one sheet;
multi-matrix products get one sheet per variant (2D, 3D, 4D, …).

Run:
    python3 scripts/export_salesqueze.py
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / 'export' / 'salesqueze'
PRODUCTS = ROOT / 'products'

CATEGORIES = ['pergola', 'vypln', 'prislusenstvi', 'sluzba']

def discover_matrix_skus():
    """Auto-discover all SKUs that have a prices.json file on disk."""
    skus = []
    for cat in CATEGORIES:
        cat_dir = PRODUCTS / cat
        if not cat_dir.exists():
            continue
        for sku_dir in sorted(cat_dir.iterdir()):
            if not sku_dir.is_dir():
                continue
            if (sku_dir / 'prices.json').exists():
                skus.append(sku_dir.name)
    return skus

HEADER_FONT = Font(name='Arial', size=10, bold=True)
HEADER_FILL = PatternFill('solid', start_color='F3F4F6')
BODY_FONT = Font(name='Arial', size=10)

def find_prices_file(sku):
    for cat in CATEGORIES:
        p = PRODUCTS / cat / sku / 'prices.json'
        if p.exists():
            return p
    return None

def write_matrix(ws, hvals, vvals, prices):
    """Write one matrix to a sheet: row 1 = widths, col A = depths, cells = prices."""
    # Column headers (widths)
    for ci, w in enumerate(hvals, 2):
        cell = ws.cell(row=1, column=ci, value=w)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Row headers (depths) + price cells
    for ri, d in enumerate(vvals, 2):
        rh = ws.cell(row=ri, column=1, value=d)
        rh.font = HEADER_FONT
        rh.fill = HEADER_FILL
        rh.alignment = Alignment(horizontal='center')
        for ci in range(len(hvals)):
            v = prices[ri - 2][ci]
            if v in (None, 0):
                continue  # skip "not available"
            cell = ws.cell(row=ri, column=ci + 2, value=v)
            cell.font = BODY_FONT
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')

    # Column widths + freeze panes
    ws.column_dimensions['A'].width = 10
    for ci in range(2, len(hvals) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = 'B2'

def export_sku(sku):
    prices_file = find_prices_file(sku)
    if not prices_file:
        print(f'  SKIP {sku}: no prices.json found')
        return False

    data = json.loads(prices_file.read_text())
    wb = Workbook()
    wb.remove(wb.active)

    if data.get('type') == 'matrix':
        ws = wb.create_sheet('prices')
        write_matrix(ws, data['horizontal_values'], data['vertical_values'], data['prices'])
        sheets_info = '1 sheet'
    elif data.get('type') == 'multi-matrix':
        for variant in data['variants']:
            label = str(variant.get('label', ''))[:31] or 'variant'
            ws = wb.create_sheet(label)
            write_matrix(ws, variant['horizontal_values'], variant['vertical_values'], variant['prices'])
        sheets_info = f'{len(data["variants"])} sheets'
    else:
        print(f'  SKIP {sku}: unknown type "{data.get("type")}"')
        return False

    out = OUT_DIR / f'{sku}.xlsx'
    wb.save(out)
    print(f'  WROTE  export/salesqueze/{sku}.xlsx  ({sheets_info})')
    return True

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    skus = discover_matrix_skus()
    if not skus:
        print('No products with prices.json found.')
        return
    written = 0
    for sku in skus:
        if export_sku(sku):
            written += 1
    print(f'\nDone. {written}/{len(skus)} files in export/salesqueze/')

if __name__ == '__main__':
    main()
