#!/usr/bin/env python3
"""Bootstrap pricelist.xlsx from pricelist.json.

Schema-agnostic: auto-discovers ALL fields from the first item in
pricelist.json. Adding new fields to the JSON does NOT require changes
to this script — columns will appear automatically.

Formatting rules are applied by column name pattern, not by position:
  - Columns matching NUMERIC_COLUMNS → number format #,##0
  - Columns matching WRAP_COLUMNS → wrap text
  - 'standard_price' → mixed (number or text, depending on value)
  - Everything else → text format @

Run:
    python3 scripts/init_xlsx.py
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / 'pricelist.json'
OUT = ROOT / 'pricelist.xlsx'

# Formatting hints by column name.
# Columns not listed here default to text format (@).
NUMERIC_COLUMNS = {'cost', 'cost_percent', 'raynet_cost'}
WRAP_COLUMNS = {'description', 'color_summary', 'raynet_description'}
MIXED_COLUMNS = {'standard_price'}  # can be number or string

MAX_WIDTH = 50
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='374151')
BODY_FONT = Font(name='Arial', size=10)

def col_kind(name):
    if name in NUMERIC_COLUMNS:
        return 'num'
    if name in WRAP_COLUMNS:
        return 'wrap'
    if name in MIXED_COLUMNS:
        return 'mixed'
    return 'text'

def build():
    if not SRC.exists():
        raise SystemExit(f'{SRC} not found — nothing to bootstrap from')
    items = json.loads(SRC.read_text())
    if not isinstance(items, list) or len(items) == 0:
        raise SystemExit(f'{SRC} is not a non-empty array')

    # Auto-discover columns from the union of all item keys, preserving
    # insertion order of the first item (Python 3.7+ dict is ordered).
    seen = set()
    columns = []
    for item in items:
        for key in item:
            if key not in seen:
                columns.append(key)
                seen.add(key)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pricelist'

    # Header row (row 1)
    for c, key in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Data rows
    for r, item in enumerate(items, 2):
        for c, key in enumerate(columns, 1):
            v = item.get(key)
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            kind = col_kind(key)
            if kind == 'num':
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif kind == 'wrap':
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif kind == 'mixed':
                cell.number_format = '#,##0' if isinstance(v, (int, float)) else '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-width, capped
    for c, key in enumerate(columns, 1):
        max_len = len(key)
        for r in range(2, len(items) + 2):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, MAX_WIDTH)

    wb.save(OUT)
    print(f'Wrote {OUT} ({OUT.stat().st_size} bytes)')
    print(f'{len(items)} rows × {len(columns)} columns: {columns}')

if __name__ == '__main__':
    build()
